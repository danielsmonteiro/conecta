from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BLUE = Font(name="Arial", color="0000FF", size=10)
BLACK = Font(name="Arial", color="000000", size=10)
GREEN = Font(name="Arial", color="008000", size=10)
BOLD = Font(name="Arial", color="000000", size=10, bold=True)
REDB = Font(name="Arial", color="B00020", size=10, bold=True)
WHITE_BOLD = Font(name="Arial", color="FFFFFF", size=11, bold=True)
TITLE = Font(name="Arial", color="0B2D5B", size=15, bold=True)
SUB = Font(name="Arial", color="555555", size=9, italic=True)
NAVY = PatternFill("solid", fgColor="0B2D5B")
LBLUE = PatternFill("solid", fgColor="DCE6F1")
YEL = PatternFill("solid", fgColor="FFF2CC")
GREY = PatternFill("solid", fgColor="F2F2F2")
REDF = PatternFill("solid", fgColor="FCE4E4")
CUR = 'R$ #,##0;(R$ #,##0);"-"'
CUR_MM = 'R$ #,##0.00,, "mi";(R$ #,##0.00,, "mi");"-"'
PCT = '0.0%'
NUM = '#,##0;(#,##0);"-"'
DEC = '#,##0.0'
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def setc(ws, coord, val, font=BLACK, fmt=None, fill=None, align=None, border=True):
    c = ws[coord]; c.value = val; c.font = font
    if fmt: c.number_format = fmt
    if fill: c.fill = fill
    if align: c.alignment = Alignment(horizontal=align, vertical="center")
    if border: c.border = BORDER
    return c

def hdr(ws, row, col, labels):
    for i, lab in enumerate(labels):
        c = ws.cell(row=row, column=col + i, value=lab)
        c.fill = NAVY; c.font = WHITE_BOLD; c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

wb = Workbook()

# ============================================================ PREMISSAS
ws = wb.active; ws.title = "Premissas"
ws.sheet_view.showGridLines = False
setc(ws, "A1", "HealthMatch — Modelo Financeiro (pivô: contingência / gap-filling)", TITLE, border=False)
setc(ws, "A2", "Azul = premissa editável · Preto = fórmula · Verde = link. Core = cobrir o GAP (descobertas/no-shows), não gerir toda a escala.", SUB, border=False)
ws.merge_cells("A1:F1"); ws.merge_cells("A2:G2")

setc(ws, "A4", "Premissas globais", BOLD, fill=LBLUE); ws.merge_cells("A4:C4")
glob = [
    ("Plantão médico 12h (R$)", 1200, CUR),
    ("Ticket médio ponderado por classe (R$)", 900, CUR),
    ("Teto do trade / fee sobre o plantão (≤5%)", 0.02, PCT),
    ("SaaS de prontidão — Coaph (R$/mês)", 8000, CUR),
    ("Custo atual de prepostos (R$/mês)", 240000, CUR),
    ("% do custo de prepostos ligado a gaps/exceções", 0.35, PCT),
    ("% dos gaps cobertos com sobrepreço hoje", 0.35, PCT),
    ("Sobrepreço médio por cobertura emergencial (R$)", 250, CUR),
    ("% dos gaps que ficariam descobertos (penalidade)", 0.06, PCT),
    ("Penalidade média por plantão descoberto (R$)", 1200, CUR),
    ("Investimento PLENO em avaliação (R$)", 1500000, CUR),
    ("Investimento do PILOTO (R$)", 300000, CUR),
]
r = 5
for name, val, fmt in glob:
    setc(ws, f"A{r}", name, BLACK); setc(ws, f"B{r}", val, BLUE, fmt, fill=YEL); r += 1
# B5 médico, B6 ticket, B7 fee%, B8 SaaS, B9 prepostos, B10 %gapPrep, B11 %sobre, B12 vSobre, B13 %desc, B14 vPen, B15 invPleno, B16 piloto
TICKET, FEE, SAAS, PREP, PGAP = "Premissas!$B$6", "Premissas!$B$7", "Premissas!$B$8", "Premissas!$B$9", "Premissas!$B$10"
PSOBRE, VSOBRE, PDESC, VPEN, INV, PILOT = "Premissas!$B$11", "Premissas!$B$12", "Premissas!$B$13", "Premissas!$B$14", "Premissas!$B$15", "Premissas!$B$16"

setc(ws, "A18", "Cenários do gap (a validar no piloto)", BOLD, fill=LBLUE); ws.merge_cells("A18:E18")
hdr(ws, 19, 1, ["Driver", "Conservador", "Realista", "Otimista"])
setc(ws, "A20", "Gaps cobertos/mês", BLACK)
setc(ws, "B20", 400, BLUE, NUM, fill=YEL); setc(ws, "C20", 600, BLUE, NUM, fill=YEL); setc(ws, "D20", 1000, BLUE, NUM, fill=YEL)
setc(ws, "A21", "% da parcela de gap automatizável", BLACK)
setc(ws, "B21", 0.30, BLUE, PCT, fill=YEL); setc(ws, "C21", 0.45, BLUE, PCT, fill=YEL); setc(ws, "D21", 0.60, BLUE, PCT, fill=YEL)
setc(ws, "A23", "Nota: sobrepreço e penalidade são as premissas de MAIOR peso e mais incertas — o piloto existe para validá-las. Ticket ponderado reflete o mix (médico R$1.200; enf/téc caem muito).", SUB, border=False)
ws.merge_cells("A23:G24")
for col, w in zip("ABCDEFG", [42, 14, 14, 14, 6, 4, 4]):
    ws.column_dimensions[col].width = w
GAPS = {"B": "B20", "C": "C20", "D": "D20"}
AUT = {"B": "B21", "C": "C21", "D": "D21"}

# ============================================================ RECEITA HM (1 Coaph)
ws2 = wb.create_sheet("Receita HM (Coaph)")
ws2.sheet_view.showGridLines = False
setc(ws2, "A1", "Receita do HealthMatch a partir da Coaph (SaaS + fee por gap)", TITLE, border=False)
setc(ws2, "A2", "Modelo híbrido: mensalidade de prontidão (fixa) + fee por gap preenchido, capado no teto do trade (≤2%).", SUB, border=False)
ws2.merge_cells("A1:E1"); ws2.merge_cells("A2:F2")
hdr(ws2, 4, 1, ["Linha (R$/mês)", "Conservador", "Realista", "Otimista"])
setc(ws2, "A5", "Fee por gap (gaps × ticket × teto)", BLACK)
for col in "BCD":
    setc(ws2, f"{col}5", f"=Premissas!{GAPS[col]}*{TICKET}*{FEE}", GREEN, CUR)
setc(ws2, "A6", "SaaS de prontidão", BLACK)
for col in "BCD":
    setc(ws2, f"{col}6", f"={SAAS}", GREEN, CUR)
setc(ws2, "A7", "Receita HM / mês", BOLD)
for col in "BCD":
    setc(ws2, f"{col}7", f"={col}5+{col}6", BOLD, CUR, fill=GREY)
setc(ws2, "A8", "Receita HM / ano", BOLD)
for col in "BCD":
    setc(ws2, f"{col}8", f"={col}7*12", BOLD, CUR, fill=LBLUE)
for col, w in zip("ABCD", [34, 15, 15, 15]):
    ws2.column_dimensions[col].width = w
# realista monthly HM revenue = C7
HMREV_M = {"B": "'Receita HM (Coaph)'!B7", "C": "'Receita HM (Coaph)'!C7", "D": "'Receita HM (Coaph)'!D7"}

# ============================================================ ECONOMIA COAPH (defensiva)
ws3 = wb.create_sheet("Economia Coaph", 1)  # posiciona como 2ª aba (núcleo da decisão)
ws3.sheet_view.showGridLines = False
setc(ws3, "A1", "Economia da Coaph (tese defensiva) — e o veredito de payback", TITLE, border=False)
setc(ws3, "A2", "Âncora: parcela do custo de prepostos ligada a gaps/exceções. Honesto: a defensiva sozinha NÃO paga R$1,5mi rápido.", SUB, border=False)
ws3.merge_cells("A1:E1"); ws3.merge_cells("A2:F2")
hdr(ws3, 4, 1, ["Economia (R$/mês)", "Conservador", "Realista", "Otimista"])
setc(ws3, "A5", "1) Tempo de preposto no gap (parcela × autom.%)", BLACK)
for col in "BCD":
    setc(ws3, f"{col}5", f"=({PREP}*{PGAP})*Premissas!{AUT[col]}", BLACK, CUR)
setc(ws3, "A6", "2) Sobrepreço emergencial evitado (gaps×%×R$)", BLACK)
for col in "BCD":
    setc(ws3, f"{col}6", f"=Premissas!{GAPS[col]}*{PSOBRE}*{VSOBRE}", BLACK, CUR)
setc(ws3, "A7", "3) Penalidade evitada (gaps×%desc×R$)", BLACK)
for col in "BCD":
    setc(ws3, f"{col}7", f"=Premissas!{GAPS[col]}*{PDESC}*{VPEN}", BLACK, CUR)
setc(ws3, "A8", "Economia BRUTA total / mês", BOLD)
for col in "BCD":
    setc(ws3, f"{col}8", f"=SUM({col}5:{col}7)", BOLD, CUR, fill=GREY)
setc(ws3, "A9", "(–) Custo pago ao HealthMatch (SaaS+fee)", BLACK)
for col in "BCD":
    setc(ws3, f"{col}9", f"=-{HMREV_M[col]}", GREEN, CUR)
setc(ws3, "A10", "Economia LÍQUIDA da Coaph / mês", BOLD)
for col in "BCD":
    setc(ws3, f"{col}10", f"={col}8+{col}9", BOLD, CUR, fill=LBLUE)
setc(ws3, "A11", "Economia líquida / ano", BOLD)
for col in "BCD":
    setc(ws3, f"{col}11", f"={col}10*12", BOLD, CUR, fill=LBLUE)
setc(ws3, "A13", "Payback (meses) sobre a economia líquida", BOLD, fill=LBLUE); ws3.merge_cells("A13:D13")
setc(ws3, "A14", "PILOTO (R$ 300k) — o ask atual", BOLD)
for col in "BCD":
    setc(ws3, f"{col}14", f"={PILOT}/{col}10", BOLD, DEC, fill=GREY)
setc(ws3, "A15", "Rodada plena (R$ 1,5 mi) — futura/ofensiva", BLACK)
for col in "BCD":
    setc(ws3, f"{col}15", f"={INV}/{col}10", BLACK, DEC)
setc(ws3, "A16", "Floor honesto: só prepostos paga R$1,5mi em (meses)", REDB)
for col in "BCD":
    setc(ws3, f"{col}16", f"={INV}/{col}5", REDB, DEC, fill=REDF)
setc(ws3, "A18", "Leitura: o PILOTO (R$300k) se paga em poucos meses mesmo no conservador. A economia depende sobretudo de sobrepreço+penalidade (linhas 2-3) — premissas que o piloto valida. Como CLIENTE, a Coaph é líquida-positiva todo mês.", SUB, border=False)
ws3.merge_cells("A18:F19")
for col, w in zip("ABCD", [44, 14, 14, 14]):
    ws3.column_dimensions[col].width = w

# ============================================================ RECEITA INCREMENTAL COAPH (vagas ociosas)
ws3b = wb.create_sheet("Receita Incremental Coaph", 2)  # 3ª aba — a alavanca de receita
ws3b.sheet_view.showGridLines = False
setc(ws3b, "A1", "Receita incremental da Coaph — preencher vagas ociosas", TITLE, border=False)
setc(ws3b, "A2", "A Coaph tem contratos que não executa por falta de profissionais. A plataforma ajuda a preencher parte das vagas ociosas → margem que hoje fica na mesa. 🔹 feeling do fundador, sem dados — células editáveis.", SUB, border=False)
ws3b.merge_cells("A1:F1"); ws3b.merge_cells("A2:G2")
setc(ws3b, "A4", "Premissas (editáveis)", BOLD, fill=LBLUE); ws3b.merge_cells("A4:C4")
inp = [("Vagas contratadas/mês", 12000, NUM), ("Vagas preenchidas hoje/mês", 8000, NUM),
       ("Valor médio da vaga (ponderado, R$)", 900, CUR), ("Margem administrativa da Coaph (<5%)", 0.04, PCT)]
r = 5
for n, v, f in inp:
    setc(ws3b, f"A{r}", n, BLACK); setc(ws3b, f"B{r}", v, BLUE, f, fill=YEL); r += 1
setc(ws3b, "A9", "Vagas ociosas/mês", BOLD); setc(ws3b, "B9", "=B5-B6", BOLD, NUM, fill=GREY)
setc(ws3b, "A10", "Taxa de ociosidade", BLACK); setc(ws3b, "B10", "=B9/B5", BLACK, PCT)
setc(ws3b, "A11", 'GMV ocioso/mês ("dinheiro na mesa")', BOLD); setc(ws3b, "B11", "=B9*B7", BOLD, CUR, fill=GREY)
setc(ws3b, "A12", "Margem total na mesa/mês (se preenchesse tudo)", BLACK); setc(ws3b, "B12", "=B11*B8", BLACK, CUR)
setc(ws3b, "A14", "Cenários de captura pela plataforma", BOLD, fill=LBLUE); ws3b.merge_cells("A14:E14")
hdr(ws3b, 15, 1, ["Métrica", "Conservador", "Realista", "Otimista"])
setc(ws3b, "A16", "% das ociosas preenchidas via plataforma", BLACK)
setc(ws3b, "B16", 0.15, BLUE, PCT, fill=YEL); setc(ws3b, "C16", 0.25, BLUE, PCT, fill=YEL); setc(ws3b, "D16", 0.40, BLUE, PCT, fill=YEL)
setc(ws3b, "A17", "Vagas recuperadas/mês", BLACK)
for col in "BCD": setc(ws3b, f"{col}17", f"=$B$9*{col}16", BLACK, NUM)
setc(ws3b, "A18", "GMV recuperado/mês", BLACK)
for col in "BCD": setc(ws3b, f"{col}18", f"={col}17*$B$7", BLACK, CUR)
setc(ws3b, "A19", "Receita incremental Coaph (margem) / mês", BOLD)
for col in "BCD": setc(ws3b, f"{col}19", f"={col}18*$B$8", BOLD, CUR, fill=LBLUE)
setc(ws3b, "A20", "Receita incremental Coaph / ano", BOLD)
for col in "BCD": setc(ws3b, f"{col}20", f"={col}19*12", BOLD, CUR)
setc(ws3b, "A22", "Benefício total Coaph/mês (economia bruta + receita incremental) — realista", BOLD, fill=LBLUE)
setc(ws3b, "B22", "='Economia Coaph'!C8+C19", BOLD, CUR, fill=LBLUE)
setc(ws3b, "A24", "Honesto: por ser cooperativa de margem fina (<5%), o ganho próprio em R$ é modesto, mas o GMV na mesa é grande e a sub-execução crônica ameaça a RENOVAÇÃO dos contratos — esse é o maior valor (estratégico). Soma-se ainda mais trabalho/renda aos cooperados.", SUB, border=False)
ws3b.merge_cells("A24:G26")
for col, w in zip("ABCDEFG", [46, 14, 14, 14, 6, 4, 4]): ws3b.column_dimensions[col].width = w

# ============================================================ RECEITA HM (escala — ofensiva)
ws4 = wb.create_sheet("Receita HM (escala)")
ws4.sheet_view.showGridLines = False
setc(ws4, "A1", "Receita do HealthMatch em escala (tese ofensiva — upside)", TITLE, border=False)
setc(ws4, "A2", "Receita por cooperativa ≈ realista da aba Receita HM. Negócio de SaaS de contingência, modesto e previsível — não marketplace de take-rate gordo.", SUB, border=False)
ws4.merge_cells("A1:E1"); ws4.merge_cells("A2:F2")
hdr(ws4, 4, 1, ["", "Ano 1", "Ano 2", "Ano 3"])
setc(ws4, "A5", "Cooperativas/clientes ativos (média)", BLACK)
setc(ws4, "B5", 1, BLUE, NUM, fill=YEL); setc(ws4, "C5", 5, BLUE, NUM, fill=YEL); setc(ws4, "D5", 15, BLUE, NUM, fill=YEL)
setc(ws4, "A6", "Receita/cliente/ano (realista)", BLACK)
for col in "BCD":
    setc(ws4, f"{col}6", "='Receita HM (Coaph)'!C8", GREEN, CUR)
setc(ws4, "A7", "Receita HM total/ano", BOLD)
for col in "BCD":
    setc(ws4, f"{col}7", f"={col}5*{col}6", BOLD, CUR, fill=LBLUE)
setc(ws4, "A9", "Premissa de ramp-up (clientes) e ticket idêntico ao realista; ajuste conforme pipeline. Receita Ano 3 modesta vs. marketplace anterior — coerente com margem fina.", SUB, border=False)
ws4.merge_cells("A9:F10")
for col, w in zip("ABCD", [34, 15, 15, 15]):
    ws4.column_dimensions[col].width = w

# ============================================================ SENSIBILIDADE
ws5 = wb.create_sheet("Sensibilidade")
ws5.sheet_view.showGridLines = False
setc(ws5, "A1", "Sensibilidade — payback (meses) da economia BRUTA vs gaps/mês × automatizável%", TITLE, border=False)
setc(ws5, "A2", "Mostra que mesmo combinações otimistas raramente pagam R$1,5mi em <30 meses só pela defensiva.", SUB, border=False)
ws5.merge_cells("A1:G1"); ws5.merge_cells("A2:G2")
gaps_axis = [300, 500, 700, 1000, 1500]
aut_axis = [0.30, 0.45, 0.60, 0.75]
setc(ws5, "A4", "gaps/mês ↓  |  automatizável% →", BOLD, fill=LBLUE)
for j, a in enumerate(aut_axis):
    setc(ws5, f"{chr(66+j)}4", a, WHITE_BOLD, PCT, fill=NAVY, align="center")
for i, g in enumerate(gaps_axis):
    rr = 5 + i
    setc(ws5, f"A{rr}", g, BOLD, NUM, fill=LBLUE, align="center")
    for j, a in enumerate(aut_axis):
        # economia bruta = parcela(prep*%gap) * aut ; payback = INV / economia
        col = chr(66 + j)
        setc(ws5, f"{col}{rr}", f"={INV}/(({PREP}*{PGAP})*{a})", BLACK, DEC, align="center")
setc(ws5, "A11", "Leitura: a parcela de prepostos é fixa (não depende do nº de gaps), por isso a defensiva tem teto de economia. Volume de gap importa para a RECEITA do HM e para o valor de cobertura, não para o payback via prepostos.", SUB, border=False)
ws5.merge_cells("A11:G12")
for col, w in zip("ABCDEFG", [26, 12, 12, 12, 12, 4, 4]):
    ws5.column_dimensions[col].width = w

wb.save("/Users/daniel/Projects/HealthMatch/business-plan/HealthMatch_Modelo_Financeiro.xlsx")
print("saved")
